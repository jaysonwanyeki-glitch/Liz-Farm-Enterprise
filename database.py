"""
Liz Farm Enterprise - SQLite Database Layer
Handles all persistence for the farm management app.
"""

import sqlite3
import pandas as pd
from datetime import datetime, timedelta
import random
import os

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "liz_farm.db")


def get_connection():
    """Get a SQLite connection with row factory."""
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def init_db():
    """Create all tables and seed data if the DB is fresh."""
    conn = get_connection()
    cursor = conn.cursor()

    # Check if already initialized — require ALL tables, not just farm_overview
    REQUIRED_TABLES = {
        "farm_overview", "inventory", "orchard", "orange_harvest",
        "cereal_shop", "cereal_inventory", "cereal_daily_sales",
        "employees", "employee_payments", "egg_production", "egg_sales",
        "fruit_production", "chicks", "pet_feeding", "mortality",
        "cat_menu", "finance", "feed_inventory", "tasks", "weather",
    }
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
    existing = {row[0] for row in cursor.fetchall()}
    if REQUIRED_TABLES.issubset(existing):
        conn.close()
        return
    # Some tables missing — drop everything and recreate fresh
    for tbl in list(existing):
        cursor.execute(f"DROP TABLE IF EXISTS [{tbl}]")
    cursor.execute("DROP TABLE IF EXISTS sqlite_sequence")
    conn.commit()

    # ── 1. farm_overview ──────────────────────────────────
    cursor.execute("""
    CREATE TABLE farm_overview (
        id          INTEGER PRIMARY KEY DEFAULT 1,
        total_land  REAL,
        owner       TEXT,
        location    TEXT,
        established INTEGER,
        orange_trees INTEGER,
        fruiting_trees INTEGER
    )""")
    cursor.execute("""
    INSERT INTO farm_overview (id, total_land, owner, location, established, orange_trees, fruiting_trees)
    VALUES (1, 2.95, 'Liz', 'Kenya', 2020, 45, 38)
    """)

    # ── 2. inventory ──────────────────────────────────────
    cursor.execute("""
    CREATE TABLE inventory (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        category        TEXT,
        male            INTEGER,
        female          INTEGER,
        total           INTEGER,
        age_months      INTEGER,
        health_status   TEXT,
        location        TEXT,
        notes           TEXT
    )""")
    inventory_rows = [
        ('🐔 Kienyeji (Adult)', 1, 7, 8, 12, '✅ Excellent', 'Free Range', 'Laying well'),
        ('🐔 Kienyeji (Pullets)', 0, 15, 15, 4, '✅ Excellent', 'Brooder', 'Growing strong'),
        ('🦆 Ducks', 1, 1, 2, 6, '✅ Excellent', 'Shed A', 'Good pair'),
        ('🐣 Chicks (New)', 0, 7, 7, 1, '✅ Excellent', 'Brooder', 'New batch'),
        ('🐱 Cats', 3, 4, 7, 24, '✅ Excellent', 'House', 'Pest control'),
        ('🐕 Dogs', 1, 1, 2, 18, '✅ Excellent', 'Kennel', 'Guard dogs'),
    ]
    cursor.executemany("INSERT INTO inventory (category,male,female,total,age_months,health_status,location,notes) VALUES (?,?,?,?,?,?,?,?)", inventory_rows)

    # ── 3. orchard ────────────────────────────────────────
    cursor.execute("""
    CREATE TABLE orchard (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        date            TEXT,
        tree_id         TEXT,
        variety         TEXT,
        age_years       INTEGER,
        fruiting        TEXT,
        harvest_kg      INTEGER,
        quality         TEXT,
        fertilizer_used TEXT,
        pest_control    TEXT,
        next_harvest    TEXT,
        notes           TEXT
    )""")
    for i in range(12):
        cursor.execute("""
        INSERT INTO orchard (date,tree_id,variety,age_years,fruiting,harvest_kg,quality,fertilizer_used,pest_control,next_harvest,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (
            (datetime.now().date() - timedelta(days=i*7)).strftime('%Y-%m-%d'),
            f'OR-{i+1:03d}', 'Valencia', random.randint(3, 7), 'Yes',
            random.randint(20, 80),
            random.choice(['Grade A', 'Grade A', 'Grade B']),
            'Organic', 'Natural',
            (datetime.now().date() + timedelta(days=random.randint(20, 60))).strftime('%Y-%m-%d'),
            ''
        ))

    # ── 4. orange_harvest ─────────────────────────────────
    cursor.execute("""
    CREATE TABLE orange_harvest (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        date                TEXT,
        quantity_kg         INTEGER,
        grade_a_kg          INTEGER,
        grade_b_kg          INTEGER,
        selling_price_kes   INTEGER,
        total_revenue_kes   INTEGER,
        harvested_by        TEXT,
        buyer               TEXT,
        payment_received    TEXT,
        notes               TEXT
    )""")
    for i in range(10):
        qty = random.randint(10, 50)
        grade_a = int(qty * random.uniform(0.6, 0.9))
        cursor.execute("""
        INSERT INTO orange_harvest (date,quantity_kg,grade_a_kg,grade_b_kg,selling_price_kes,total_revenue_kes,harvested_by,buyer,payment_received,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            (datetime.now().date() - timedelta(days=i*3)).strftime('%Y-%m-%d'),
            qty, grade_a, qty - grade_a, 200, qty * 200,
            'John', 'Local Market',
            random.choice(['✅ Yes', '⏳ Pending'] + ['✅ Yes']*3), ''
        ))

    # ── 5. cereal_shop ────────────────────────────────────
    cursor.execute("""
    CREATE TABLE cereal_shop (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        date                TEXT,
        cereal_type         TEXT,
        quantity_kg         INTEGER,
        buying_price_kes    INTEGER,
        selling_price_kes   INTEGER,
        total_cost_kes      INTEGER,
        total_revenue_kes   INTEGER,
        profit_loss_kes     INTEGER,
        sold_by             TEXT,
        payment_method      TEXT,
        customer_type       TEXT,
        notes               TEXT
    )""")
    cereals = ['🌾 Maize', '🌾 Beans', '🌾 Rice', '🌾 Wheat', '🌾 Millet', '🌾 Sorghum', '🌾 Green Grams']
    for i in range(20):
        cereal = random.choice(cereals)
        qty = random.randint(10, 100)
        bp = random.randint(50, 150)
        sp = bp + random.randint(30, 80)
        cursor.execute("""
        INSERT INTO cereal_shop (date,cereal_type,quantity_kg,buying_price_kes,selling_price_kes,total_cost_kes,total_revenue_kes,profit_loss_kes,sold_by,payment_method,customer_type,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            (datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d'),
            cereal, qty, bp, sp,
            qty * bp, qty * sp, qty * (sp - bp),
            'Grace',
            random.choice(['Cash', 'M-Pesa', 'Bank']),
            random.choice(['Retail', 'Wholesale']), ''
        ))

    # ── 5b. cereal_inventory ──────────────────────────────
    cursor.execute("""
    CREATE TABLE cereal_inventory (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        cereal_type         TEXT,
        stock_kg            INTEGER,
        buying_price_kes    INTEGER,
        selling_price_kes   INTEGER,
        min_stock_kg        INTEGER,
        max_stock_kg        INTEGER,
        supplier            TEXT,
        last_restocked      TEXT,
        status              TEXT,
        notes               TEXT
    )""")
    cereal_inv_rows = [
        ('🌾 Maize', 500, 65, 110, 100, 1000, 'Nakuru Millers', (datetime.now().date() - timedelta(days=2)).strftime('%Y-%m-%d'), '✅ In Stock', ''),
        ('🌾 Beans', 200, 90, 155, 50, 500, 'Nakuru Millers', (datetime.now().date() - timedelta(days=3)).strftime('%Y-%m-%d'), '✅ In Stock', ''),
        ('🌾 Rice', 150, 120, 190, 40, 400, 'Mombasa Supplies', (datetime.now().date() - timedelta(days=1)).strftime('%Y-%m-%d'), '✅ In Stock', ''),
        ('🌾 Wheat', 180, 80, 135, 50, 400, 'Nakuru Millers', (datetime.now().date() - timedelta(days=4)).strftime('%Y-%m-%d'), '✅ In Stock', ''),
        ('🌾 Millet', 100, 70, 120, 30, 300, 'Local Market', (datetime.now().date() - timedelta(days=5)).strftime('%Y-%m-%d'), '✅ In Stock', ''),
        ('🌾 Sorghum', 120, 60, 105, 30, 300, 'Local Market', (datetime.now().date() - timedelta(days=2)).strftime('%Y-%m-%d'), '✅ In Stock', ''),
        ('🌾 Green Grams', 80, 100, 170, 25, 250, 'Mombasa Supplies', (datetime.now().date() - timedelta(days=6)).strftime('%Y-%m-%d'), '✅ In Stock', ''),
    ]
    cursor.executemany("INSERT INTO cereal_inventory (cereal_type,stock_kg,buying_price_kes,selling_price_kes,min_stock_kg,max_stock_kg,supplier,last_restocked,status,notes) VALUES (?,?,?,?,?,?,?,?,?,?)", cereal_inv_rows)

    # ── 5c. cereal_daily_sales ─────────────────────────────
    cursor.execute("""
    CREATE TABLE cereal_daily_sales (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        date                TEXT,
        total_items_sold    INTEGER,
        total_kg_sold       INTEGER,
        total_cost_kes      INTEGER,
        total_revenue_kes   INTEGER,
        profit_kes          INTEGER,
        cash_sales_kes      INTEGER,
        mpesa_sales_kes     INTEGER,
        bank_sales_kes      INTEGER,
        customers_served    INTEGER,
        sold_by             TEXT,
        notes               TEXT
    )""")
    for i in range(30):
        items = random.randint(3, 12)
        kg = random.randint(40, 250)
        cost = kg * random.randint(60, 120)
        rev = kg * random.randint(100, 180)
        cash_pct = random.uniform(0.3, 0.7)
        mpesa_pct = random.uniform(0.2, 0.5)
        bank_pct = 1 - cash_pct - mpesa_pct
        cursor.execute("""
        INSERT INTO cereal_daily_sales (date,total_items_sold,total_kg_sold,total_cost_kes,total_revenue_kes,profit_kes,cash_sales_kes,mpesa_sales_kes,bank_sales_kes,customers_served,sold_by,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            (datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d'),
            items, kg, cost, rev, rev - cost,
            int(rev * cash_pct), int(rev * mpesa_pct), int(rev * bank_pct),
            random.randint(5, 25), 'Grace', ''
        ))

    # ── 6. employees ──────────────────────────────────────
    cursor.execute("""
    CREATE TABLE employees (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        employee_name   TEXT,
        role            TEXT,
        work_location   TEXT,
        start_date      TEXT,
        salary_kes      INTEGER,
        payment_method  TEXT,
        phone           TEXT,
        status          TEXT,
        notes           TEXT
    )""")
    cursor.executemany("INSERT INTO employees (employee_name,role,work_location,start_date,salary_kes,payment_method,phone,status,notes) VALUES (?,?,?,?,?,?,?,?,?)", [
        ('John Mwangi', 'Farm Worker', 'Shamba (Farm)', (datetime.now().date() - timedelta(days=180)).strftime('%Y-%m-%d'), 20000, 'M-Pesa', '0712345678', '✅ Active', 'Experienced farmer'),
        ('Grace Akinyi', 'Cereal Shop Attendant', 'Cereal Shop', (datetime.now().date() - timedelta(days=150)).strftime('%Y-%m-%d'), 18000, 'M-Pesa', '0723456789', '✅ Active', 'Good with customers'),
    ])

    # ── 7. employee_payments ──────────────────────────────
    cursor.execute("""
    CREATE TABLE employee_payments (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        month           TEXT,
        employee        TEXT,
        amount_kes      INTEGER,
        date_paid       TEXT,
        payment_method  TEXT,
        status          TEXT
    )""")
    for i in range(6):
        month = (datetime.now().date() - timedelta(days=30*i)).strftime('%Y-%m')
        cursor.execute("INSERT INTO employee_payments (month,employee,amount_kes,date_paid,payment_method,status) VALUES (?,?,?,?,?,?)",
            (month, 'John Mwangi', 20000, (datetime.now().date() - timedelta(days=30*i+5)).strftime('%Y-%m-%d'), 'M-Pesa', '✅ Paid'))
        cursor.execute("INSERT INTO employee_payments (month,employee,amount_kes,date_paid,payment_method,status) VALUES (?,?,?,?,?,?)",
            (month, 'Grace Akinyi', 18000, (datetime.now().date() - timedelta(days=30*i+7)).strftime('%Y-%m-%d'), 'M-Pesa', '✅ Paid'))

    # ── 8. egg_production ─────────────────────────────────
    cursor.execute("""
    CREATE TABLE egg_production (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        date            TEXT,
        day_name        TEXT,
        month_name      TEXT,
        trays           INTEGER,
        eggs_per_tray   INTEGER,
        total_eggs      INTEGER,
        sellable        INTEGER,
        cracked         INTEGER,
        quality         TEXT,
        notes           TEXT
    )""")
    for i in range(30):
        d = datetime.now().date() + timedelta(days=i)
        trays = random.randint(2, 6)
        total = trays * 30
        cracked = random.randint(0, 5)
        cursor.execute("""
        INSERT INTO egg_production (date,day_name,month_name,trays,eggs_per_tray,total_eggs,sellable,cracked,quality,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            d.strftime('%Y-%m-%d'), d.strftime('%A'), d.strftime('%B'),
            trays, 30, total, total - cracked, cracked,
            random.choice(['Grade A', 'Grade A', 'Grade A', 'Grade B']), ''
        ))

    # ── 9. egg_sales ──────────────────────────────────────
    cursor.execute("""
    CREATE TABLE egg_sales (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        date                TEXT,
        trays_sold          INTEGER,
        price_per_tray_kes  INTEGER,
        total_revenue_kes   INTEGER,
        customer            TEXT,
        payment_status      TEXT,
        notes               TEXT
    )""")
    for i in range(15):
        trays = random.randint(1, 5)
        cursor.execute("""
        INSERT INTO egg_sales (date,trays_sold,price_per_tray_kes,total_revenue_kes,customer,payment_status,notes)
        VALUES (?,?,?,?,?,?,?)
        """, (
            (datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d'),
            trays, 500, trays * 500, 'Local Market',
            random.choice(['✅ Paid', '✅ Paid', '⏳ Pending']), ''
        ))

    # ── 10. fruit_production ──────────────────────────────
    cursor.execute("""
    CREATE TABLE fruit_production (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        date                TEXT,
        fruit_type          TEXT,
        quantity            INTEGER,
        unit                TEXT,
        quality             TEXT,
        harvested_by        TEXT,
        selling_price_kes   INTEGER,
        total_revenue_kes   INTEGER,
        notes               TEXT
    )""")
    fruits = ['🍊 Oranges', '🥥 Coconuts', '🍍 Pawpaw', '🌽 Maize']
    for i in range(14):
        fruit = random.choice(fruits)
        qty = random.randint(10, 60)
        sp = random.randint(150, 300)
        cursor.execute("""
        INSERT INTO fruit_production (date,fruit_type,quantity,unit,quality,harvested_by,selling_price_kes,total_revenue_kes,notes)
        VALUES (?,?,?,?,?,?,?,?,?)
        """, (
            (datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d'),
            fruit, qty, 'kg',
            random.choice(['Grade A', 'Grade A', 'Grade B']),
            'John', sp, qty * sp, ''
        ))

    # ── 11. chicks ────────────────────────────────────────
    cursor.execute("""
    CREATE TABLE chicks (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        hatch_date      TEXT,
        breed           TEXT,
        hatched         INTEGER,
        survival        INTEGER,
        survival_pct    INTEGER,
        location        TEXT,
        vaccinated      TEXT,
        notes           TEXT
    )""")
    for i, days_ago in enumerate([7, 5, 3]):
        cursor.execute("INSERT INTO chicks (hatch_date,breed,hatched,survival,survival_pct,location,vaccinated,notes) VALUES (?,?,?,?,?,?,?,?)",
            ((datetime.now().date() - timedelta(days=days_ago)).strftime('%Y-%m-%d'),
             'Kienyeji', [3, 2, 2][i], [3, 2, 2][i], 100,
             ['Brooder 1', 'Brooder 2', 'Brooder 1'][i],
             ['✅ Yes', '✅ Yes', '⏳ Pending'][i],
             ['Healthy', 'Active', 'Growing well'][i]))

    # ── 12. pet_feeding ───────────────────────────────────
    cursor.execute("""
    CREATE TABLE pet_feeding (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        date            TEXT,
        pet_type        TEXT,
        meal            TEXT,
        food_type       TEXT,
        quantity_kg     REAL,
        cost_per_kg_kes INTEGER,
        total_cost_kes  INTEGER,
        fed_by          TEXT,
        animals_fed     TEXT,
        notes           TEXT
    )""")
    today = datetime.now().date()
    for i in range(10):
        qty = round(random.uniform(0.5, 2.0), 1)
        cpk = random.randint(150, 350)
        cursor.execute("""
        INSERT INTO pet_feeding (date,pet_type,meal,food_type,quantity_kg,cost_per_kg_kes,total_cost_kes,fed_by,animals_fed,notes)
        VALUES (?,?,?,?,?,?,?,?,?,?)
        """, (
            (today + timedelta(days=i//2)).strftime('%Y-%m-%d'),
            random.choice(['🐱 Cats', '🐕 Dogs']),
            ['Morning', 'Evening'][i % 2],
            random.choice(['🐟 Fish + Rice', '🍗 Chicken + Liver', '🥩 Meat + Vegetables', '🥚 Eggs + Rice']),
            qty, cpk, round(qty * cpk), 'Liz',
            random.choice(['7/7', '2/2']), ''
        ))

    # ── 13. mortality ─────────────────────────────────────
    cursor.execute("""
    CREATE TABLE mortality (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        date            TEXT,
        species         TEXT,
        age             TEXT,
        lost            INTEGER,
        reason          TEXT,
        action_taken    TEXT,
        prevention      TEXT,
        severity        TEXT
    )""")
    cursor.execute("INSERT INTO mortality (date,species,age,lost,reason,action_taken,prevention,severity) VALUES (?,?,?,?,?,?,?,?)",
        (datetime.now().date().strftime('%Y-%m-%d'), 'None', '0', 0, 'No losses recorded', 'None', 'Monitoring', 'Low'))

    # ── 14. cat_menu ──────────────────────────────────────
    cursor.execute("""
    CREATE TABLE cat_menu (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        date            TEXT,
        day_name        TEXT,
        meal            TEXT,
        ingredients     TEXT,
        portion_cups    REAL,
        schedule        TEXT,
        cats_fed        TEXT,
        notes           TEXT
    )""")
    meals_list = ['Breakfast', 'Lunch', 'Dinner', 'Evening Treat']
    ingredients_list = ['Dry kibble + warm milk', 'Fish + rice + vegetables', 'Chicken + liver + broth', 'Catnip + tuna + egg']
    portions = [0.5, 0.75, 0.5, 0.25]
    schedules = ['6:00 AM', '12:00 PM', '6:00 PM', '9:00 PM']
    for i in range(8):
        d = today + timedelta(days=i//4)
        cursor.execute("INSERT INTO cat_menu (date,day_name,meal,ingredients,portion_cups,schedule,cats_fed,notes) VALUES (?,?,?,?,?,?,?,?)",
            (d.strftime('%Y-%m-%d'), d.strftime('%A'), meals_list[i % 4],
             ingredients_list[i % 4], portions[i % 4], schedules[i % 4], '7/7', ''))

    # ── 15. finance ───────────────────────────────────────
    cursor.execute("""
    CREATE TABLE finance (
        id                  INTEGER PRIMARY KEY AUTOINCREMENT,
        date                TEXT,
        source_expense      TEXT,
        type                TEXT,
        amount_kes          INTEGER,
        category            TEXT,
        payment_method      TEXT,
        notes               TEXT
    )""")
    fin_rows = [
        ('Egg Sales', 'Income', 5000, 'Poultry', 'M-Pesa'),
        ('Orange Sales', 'Income', 3000, 'Orchard', 'Cash'),
        ('Cereal Sales', 'Income', 8000, 'Cereal Shop', 'M-Pesa'),
        ('Chicken Feed', 'Expense', 2000, 'Poultry', 'Cash'),
        ('Employee Salary', 'Expense', 38000, 'Labor', 'Bank'),
        ('Vaccinations', 'Expense', 1500, 'Health', 'Cash'),
        ('Fertilizer', 'Expense', 3000, 'Farming', 'Cash'),
        ('Transport', 'Expense', 2000, 'Logistics', 'M-Pesa'),
        ('Pet Food', 'Expense', 1000, 'Pets', 'Cash'),
        ('Miscellaneous', 'Expense', 1500, 'Other', 'Cash'),
    ]
    for i, (src, typ, amt, cat, pm) in enumerate(fin_rows):
        cursor.execute("INSERT INTO finance (date,source_expense,type,amount_kes,category,payment_method,notes) VALUES (?,?,?,?,?,?,?)",
            ((datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d'),
             src, typ, amt, cat, pm, ''))

    # ── 16. feed_inventory ────────────────────────────────
    cursor.execute("""
    CREATE TABLE feed_inventory (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        feed_type       TEXT,
        quantity_kg     INTEGER,
        cost_per_kg_kes INTEGER,
        supplier        TEXT,
        last_ordered    TEXT,
        reorder_level   INTEGER,
        status          TEXT
    )""")
    feed_rows = [
        ('🐔 Kienyeji Mix', 150, 100, 'Local Mill', 30, '✅ In Stock'),
        ('🐔 Chick Starter', 50, 120, 'Kenya Feeds', 15, '✅ In Stock'),
        ('🦆 Duck Feed', 20, 130, 'Kenya Feeds', 10, '✅ In Stock'),
        ('🐱 Cat Food', 15, 200, 'Petshop', 5, '✅ In Stock'),
        ('🐕 Dog Food', 12, 180, 'Petshop', 5, '✅ In Stock'),
        ('🌾 Maize', 200, 80, 'Farm Store', 50, '✅ In Stock'),
    ]
    for i, (ft, qty, cpk, sup, rl, status) in enumerate(feed_rows):
        cursor.execute("INSERT INTO feed_inventory (feed_type,quantity_kg,cost_per_kg_kes,supplier,last_ordered,reorder_level,status) VALUES (?,?,?,?,?,?,?)",
            (ft, qty, cpk, sup, (datetime.now().date() - timedelta(days=i*2)).strftime('%Y-%m-%d'), rl, status))

    # ── 17. tasks ─────────────────────────────────────────
    cursor.execute("""
    CREATE TABLE tasks (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        date            TEXT,
        task            TEXT,
        priority        TEXT,
        assigned_to     TEXT,
        status          TEXT,
        category        TEXT,
        notes           TEXT
    )""")
    task_rows = [
        ('Collect eggs - Poultry', 'High', 'John', '✅ Done', 'Poultry'),
        ('Feed chickens - Poultry', 'High', 'John', '✅ Done', 'Poultry'),
        ('Water orange trees - Orchard', 'High', 'John', '✅ Done', 'Orchard'),
        ('Harvest oranges - Orchard', 'Medium', 'John', '✅ Done', 'Orchard'),
        ('Clean cereal shop - Shop', 'High', 'Grace', '✅ Done', 'Shop'),
        ('Restock cereals - Shop', 'Medium', 'Grace', '✅ Done', 'Shop'),
        ('Health check - All', 'High', 'John', '✅ Done', 'Health'),
        ('Clean chicken sheds - Poultry', 'Medium', 'John', '✅ Done', 'Maintenance'),
        ('Check pet feeding - Pets', 'Medium', 'Liz', '✅ Done', 'Pets'),
    ]
    today_str = datetime.now().date().strftime('%Y-%m-%d')
    for task, pri, who, stat, cat in task_rows:
        cursor.execute("INSERT INTO tasks (date,task,priority,assigned_to,status,category,notes) VALUES (?,?,?,?,?,?,?)",
            (today_str, task, pri, who, stat, cat, ''))

    # ── 18. weather ───────────────────────────────────────
    cursor.execute("""
    CREATE TABLE weather (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        date            TEXT,
        temperature_c   INTEGER,
        humidity_pct    INTEGER,
        weather         TEXT,
        rainfall_mm     INTEGER,
        notes           TEXT
    )""")
    weather_types = ['☀️ Sunny', '⛅ Cloudy', '🌧️ Rainy', '🌤️ Partly Cloudy']
    for i in range(7):
        cursor.execute("INSERT INTO weather (date,temperature_c,humidity_pct,weather,rainfall_mm,notes) VALUES (?,?,?,?,?,?)",
            ((datetime.now().date() - timedelta(days=i)).strftime('%Y-%m-%d'),
             random.randint(22, 32), random.randint(45, 85),
             random.choice(weather_types), random.randint(0, 20), ''))

    conn.commit()
    conn.close()


# ═══════════════════════════════════════════════════════════
#  GENERIC HELPERS
# ═══════════════════════════════════════════════════════════

def _df(sql, params=()):
    """Return a DataFrame from a SQL query."""
    conn = get_connection()
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()
    return df


def _execute(sql, params=()):
    conn = get_connection()
    conn.execute(sql, params)
    conn.commit()
    conn.close()


def _execute_many(sql, rows):
    conn = get_connection()
    conn.executemany(sql, rows)
    conn.commit()
    conn.close()


# ═══════════════════════════════════════════════════════════
#  FARM OVERVIEW
# ═══════════════════════════════════════════════════════════

def get_farm_overview():
    row = _df("SELECT * FROM farm_overview WHERE id=1").iloc[0]
    return {
        'total_land': row['total_land'],
        'owner': row['owner'],
        'location': row['location'],
        'established': int(row['established']),
        'orange_trees': int(row['orange_trees']),
        'fruiting_trees': int(row['fruiting_trees']),
    }


# ═══════════════════════════════════════════════════════════
#  GENERIC TABLE <-> DATAFRAME  (used by most sections)
# ═══════════════════════════════════════════════════════════

_TABLE_MAP = {
    'inventory_data':         ('inventory',         'id'),
    'orchard_data':           ('orchard',            'id'),
    'orange_harvest_data':    ('orange_harvest',     'id'),
    'cereal_data':            ('cereal_shop',        'id'),
    'cereal_inv_data':        ('cereal_inventory',   'id'),
    'cereal_daily_data':      ('cereal_daily_sales', 'id'),
    'employee_data':          ('employees',          'id'),
    'payments_data':          ('employee_payments',  'id'),
    'egg_data':               ('egg_production',     'id'),
    'egg_sales_data':         ('egg_sales',          'id'),
    'fruit_data':             ('fruit_production',   'id'),
    'chick_data':             ('chicks',             'id'),
    'pet_feed_data':          ('pet_feeding',        'id'),
    'mortality_data':         ('mortality',          'id'),
    'cat_menu_data':          ('cat_menu',           'id'),
    'finance_data':           ('finance',            'id'),
    'feed_data':              ('feed_inventory',     'id'),
    'tasks_data':             ('tasks',              'id'),
    'weather_data':           ('weather',            'id'),
}

# Maps session_state key -> (db table, id column, column alias map)
# alias map: streamlit column name -> db column name
_ALIAS_MAP = {
    'inventory_data': {
        'Category': 'category', 'Male': 'male', 'Female': 'female',
        'Total': 'total', 'Age (months)': 'age_months',
        'Health Status': 'health_status', 'Location': 'location', 'Notes': 'notes',
    },
    'orchard_data': {
        'Date': 'date', 'Tree ID': 'tree_id', 'Variety': 'variety',
        'Age (years)': 'age_years', 'Fruiting': 'fruiting',
        'Harvest (kg)': 'harvest_kg', 'Quality': 'quality',
        'Fertilizer Used': 'fertilizer_used', 'Pest Control': 'pest_control',
        'Next Harvest': 'next_harvest', 'Notes': 'notes',
    },
    'orange_harvest_data': {
        'Date': 'date', 'Quantity (kg)': 'quantity_kg',
        'Grade A (kg)': 'grade_a_kg', 'Grade B (kg)': 'grade_b_kg',
        'Selling Price (KES/kg)': 'selling_price_kes',
        'Total Revenue (KES)': 'total_revenue_kes',
        'Harvested By': 'harvested_by', 'Buyer': 'buyer',
        'Payment Received': 'payment_received', 'Notes': 'notes',
    },
    'cereal_data': {
        'Date': 'date', 'Cereal Type': 'cereal_type',
        'Quantity (kg)': 'quantity_kg', 'Buying Price (KES/kg)': 'buying_price_kes',
        'Selling Price (KES/kg)': 'selling_price_kes',
        'Total Cost (KES)': 'total_cost_kes', 'Total Revenue (KES)': 'total_revenue_kes',
        'Profit/Loss (KES)': 'profit_loss_kes', 'Sold By': 'sold_by',
        'Payment Method': 'payment_method', 'Customer Type': 'customer_type',
        'Notes': 'notes',
    },
    'cereal_inv_data': {
        'Cereal Type': 'cereal_type', 'Stock (kg)': 'stock_kg',
        'Buying Price (KES/kg)': 'buying_price_kes', 'Selling Price (KES/kg)': 'selling_price_kes',
        'Min Stock (kg)': 'min_stock_kg', 'Max Stock (kg)': 'max_stock_kg',
        'Supplier': 'supplier', 'Last Restocked': 'last_restocked',
        'Status': 'status', 'Notes': 'notes',
    },
    'cereal_daily_data': {
        'Date': 'date', 'Items Sold': 'total_items_sold',
        'Kg Sold': 'total_kg_sold', 'Total Cost (KES)': 'total_cost_kes',
        'Total Revenue (KES)': 'total_revenue_kes', 'Profit (KES)': 'profit_kes',
        'Cash Sales (KES)': 'cash_sales_kes', 'M-Pesa Sales (KES)': 'mpesa_sales_kes',
        'Bank Sales (KES)': 'bank_sales_kes', 'Customers': 'customers_served',
        'Sold By': 'sold_by', 'Notes': 'notes',
    },
    'employee_data': {
        'Employee Name': 'employee_name', 'Role': 'role',
        'Work Location': 'work_location', 'Start Date': 'start_date',
        'Salary (KES/month)': 'salary_kes', 'Payment Method': 'payment_method',
        'Phone': 'phone', 'Status': 'status', 'Notes': 'notes',
    },
    'payments_data': {
        'Month': 'month', 'Employee': 'employee',
        'Amount (KES)': 'amount_kes', 'Date Paid': 'date_paid',
        'Payment Method': 'payment_method', 'Status': 'status',
    },
    'egg_data': {
        'Date': 'date', 'Day': 'day_name', 'Month': 'month_name',
        'Trays': 'trays', 'Eggs per Tray': 'eggs_per_tray',
        'Total Eggs': 'total_eggs', 'Sellable': 'sellable',
        'Cracked': 'cracked', 'Quality': 'quality', 'Notes': 'notes',
    },
    'egg_sales_data': {
        'Date': 'date', 'Trays Sold': 'trays_sold',
        'Price per Tray (KES)': 'price_per_tray_kes',
        'Total Revenue (KES)': 'total_revenue_kes',
        'Customer': 'customer', 'Payment Status': 'payment_status',
        'Notes': 'notes',
    },
    'fruit_data': {
        'Date': 'date', 'Fruit Type': 'fruit_type',
        'Quantity': 'quantity', 'Unit': 'unit', 'Quality': 'quality',
        'Harvested By': 'harvested_by',
        'Selling Price (KES/kg)': 'selling_price_kes',
        'Total Revenue (KES)': 'total_revenue_kes', 'Notes': 'notes',
    },
    'chick_data': {
        'Hatch Date': 'hatch_date', 'Breed': 'breed',
        '# Hatched': 'hatched', 'Survival': 'survival',
        'Survival %': 'survival_pct', 'Location': 'location',
        'Vaccinated': 'vaccinated', 'Notes': 'notes',
    },
    'pet_feed_data': {
        'Date': 'date', 'Pet Type': 'pet_type', 'Meal': 'meal',
        'Food Type': 'food_type', 'Quantity (kg)': 'quantity_kg',
        'Cost per kg (KES)': 'cost_per_kg_kes',
        'Total Cost (KES)': 'total_cost_kes', 'Fed By': 'fed_by',
        'Animals Fed': 'animals_fed', 'Notes': 'notes',
    },
    'mortality_data': {
        'Date': 'date', 'Species': 'species', 'Age': 'age',
        'Lost': 'lost', 'Reason': 'reason',
        'Action Taken': 'action_taken', 'Prevention': 'prevention',
        'Severity': 'severity',
    },
    'cat_menu_data': {
        'Date': 'date', 'Day': 'day_name', 'Meal': 'meal',
        'Ingredients': 'ingredients', 'Portion (cups)': 'portion_cups',
        'Schedule': 'schedule', 'Cats Fed': 'cats_fed', 'Notes': 'notes',
    },
    'finance_data': {
        'Date': 'date', 'Source/Expense': 'source_expense',
        'Type': 'type', 'Amount (KES)': 'amount_kes',
        'Category': 'category', 'Payment Method': 'payment_method',
        'Notes': 'notes',
    },
    'feed_data': {
        'Feed Type': 'feed_type', 'Quantity (kg)': 'quantity_kg',
        'Cost per kg (KES)': 'cost_per_kg_kes', 'Supplier': 'supplier',
        'Last Ordered': 'last_ordered', 'Reorder Level (kg)': 'reorder_level',
        'Status': 'status',
    },
    'tasks_data': {
        'Date': 'date', 'Task': 'task', 'Priority': 'priority',
        'Assigned To': 'assigned_to', 'Status': 'status',
        'Category': 'category', 'Notes': 'notes',
    },
    'weather_data': {
        'Date': 'date', 'Temperature (C)': 'temperature_c',
        'Humidity (%)': 'humidity_pct', 'Weather': 'weather',
        'Rainfall (mm)': 'rainfall_mm', 'Notes': 'notes',
    },
}


def load_table(session_key):
    """Load a table from DB into a DataFrame with Streamlit column names."""
    if session_key == 'farm_overview':
        return get_farm_overview()
    table, id_col = _TABLE_MAP[session_key]
    alias = _ALIAS_MAP[session_key]
    db_cols = ', '.join(alias.values())
    df = _df(f"SELECT {db_cols} FROM {table} ORDER BY {id_col}")
    # Rename db -> streamlit
    inv = {v: k for k, v in alias.items()}
    df = df.rename(columns=inv)
    return df


def save_table(session_key, df):
    """Save a DataFrame back to the DB (full replace for that table)."""
    if session_key == 'farm_overview':
        _execute(
            "UPDATE farm_overview SET total_land=?, owner=?, location=?, established=?, orange_trees=?, fruiting_trees=? WHERE id=1",
            (df['total_land'], df['owner'], df['location'], int(df['established']),
             int(df['orange_trees']), int(df['fruiting_trees']))
        )
        return
    table, id_col = _TABLE_MAP[session_key]
    alias = _ALIAS_MAP[session_key]
    db_cols = [alias[c] for c in df.columns if c in alias]
    # Clear table
    _execute(f"DELETE FROM {table}")
    # Insert rows
    if len(df) > 0:
        placeholders = ','.join(['?'] * len(db_cols))
        col_names = ','.join(db_cols)
        sql = f"INSERT INTO {table} ({col_names}) VALUES ({placeholders})"
        rows = []
        for _, row in df.iterrows():
            rows.append(tuple(row[c] for c in df.columns if c in alias))
        _execute_many(sql, rows)


def load_all():
    """Load all tables into session_state. Called once at startup."""
    import streamlit as st
    from database import load_table as _lt
    keys = list(_TABLE_MAP.keys())
    for key in keys:
        data = _lt(key)
        if isinstance(data, pd.DataFrame):
            st.session_state[key] = data
    # Farm overview is a dict — convert numpy types to native Python
    ov = _lt('farm_overview')
    st.session_state.farm_overview = {k: int(v) if isinstance(v, (int,)) else float(v) if isinstance(v, (float,)) else v for k, v in ov.items()}


def st_import_data(key, df):
    """Store a DataFrame into st.session_state."""
    import streamlit as st
    st.session_state[key] = df


# ═══════════════════════════════════════════════════════════
#  BACKUP & RESTORE
# ═══════════════════════════════════════════════════════════

def backup_database():
    """Return the raw bytes of the SQLite database for download."""
    if not os.path.exists(DB_PATH):
        return None
    # Use a temp file to create a consistent snapshot via SQLite backup API
    import tempfile
    src = get_connection()
    with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
        tmp_path = tmp.name
    try:
        dst = sqlite3.connect(tmp_path)
        src.backup(dst)
        dst.close()
        with open(tmp_path, 'rb') as f:
            data = f.read()
    finally:
        src.close()
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    return data


def backup_info():
    """Return metadata about the current database."""
    if not os.path.exists(DB_PATH):
        return {'exists': False}
    size_bytes = os.path.getsize(DB_PATH)
    mod_time = datetime.fromtimestamp(os.path.getmtime(DB_PATH))
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name != 'sqlite_sequence'")
    tables = [row[0] for row in cursor.fetchall()]
    total_rows = 0
    for t in tables:
        cursor.execute(f"SELECT COUNT(*) FROM [{t}]")
        total_rows += cursor.fetchone()[0]
    conn.close()
    return {
        'exists': True,
        'size_bytes': size_bytes,
        'size_mb': round(size_bytes / (1024 * 1024), 2),
        'modified': mod_time.strftime('%Y-%m-%d %H:%M:%S'),
        'tables': len(tables),
        'total_rows': total_rows,
        'path': DB_PATH,
    }


def restore_database(file_bytes):
    """Replace the current database with uploaded bytes.
    
    Returns (success: bool, message: str).
    """
    import shutil
    import tempfile
    try:
        # Ensure file_bytes is raw bytes
        if hasattr(file_bytes, 'read'):
            file_bytes = file_bytes.read()
        
        # Write to temp file for validation
        with tempfile.NamedTemporaryFile(suffix='.db', delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        
        try:
            # Validate the uploaded file is a real SQLite database
            test_conn = sqlite3.connect(tmp_path)
            cursor = test_conn.cursor()
            cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
            tables_found = [row[0] for row in cursor.fetchall()]
            test_conn.close()
            
            if len(tables_found) < 5:
                return False, f"Invalid database: only found {len(tables_found)} tables. Expected 18+ tables."
            
            # Backup current DB before overwriting
            if os.path.exists(DB_PATH):
                backup_path = DB_PATH + '.pre_restore_backup'
                shutil.copy2(DB_PATH, backup_path)
            
            # Write the new database
            shutil.copy2(tmp_path, DB_PATH)
            
            return True, f"Restored successfully! Found {len(tables_found)} tables: {', '.join(tables_found[:5])}..."
        finally:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
    except Exception as e:
        return False, f"Restore failed: {str(e)}"


def get_table_stats():
    """Return per-table row counts for the backup/restore UI."""
    if not os.path.exists(DB_PATH):
        return []
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("""
        SELECT name FROM sqlite_master 
        WHERE type='table' AND name != 'sqlite_sequence'
        ORDER BY name
    """)
    tables = [row[0] for row in cursor.fetchall()]
    stats = []
    for t in tables:
        cursor.execute(f"SELECT COUNT(*) FROM [{t}]")
        count = cursor.fetchone()[0]
        stats.append({'table': t, 'rows': count})
    conn.close()
    return stats


# ═══════════════════════════════════════════════════════════
#  COLUMN ORDER & EXCEL FORMATTING
# ═══════════════════════════════════════════════════════════

# Canonical column order for every table (Streamlit names)
COLUMN_ORDER = {
    'inventory_data': [
        'Category', 'Male', 'Female', 'Total', 'Age (months)',
        'Health Status', 'Location', 'Notes'
    ],
    'orchard_data': [
        'Date', 'Tree ID', 'Variety', 'Age (years)', 'Fruiting',
        'Harvest (kg)', 'Quality', 'Fertilizer Used', 'Pest Control',
        'Next Harvest', 'Notes'
    ],
    'orange_harvest_data': [
        'Date', 'Quantity (kg)', 'Grade A (kg)', 'Grade B (kg)',
        'Selling Price (KES/kg)', 'Total Revenue (KES)',
        'Harvested By', 'Buyer', 'Payment Received', 'Notes'
    ],
    'cereal_data': [
        'Date', 'Cereal Type', 'Quantity (kg)', 'Buying Price (KES/kg)',
        'Selling Price (KES/kg)', 'Total Cost (KES)',
        'Total Revenue (KES)', 'Profit/Loss (KES)', 'Sold By',
        'Payment Method', 'Customer Type', 'Notes'
    ],
    'cereal_inv_data': [
        'Cereal Type', 'Stock (kg)', 'Buying Price (KES/kg)',
        'Selling Price (KES/kg)', 'Min Stock (kg)', 'Max Stock (kg)',
        'Supplier', 'Last Restocked', 'Status', 'Notes'
    ],
    'cereal_daily_data': [
        'Date', 'Items Sold', 'Kg Sold', 'Total Cost (KES)',
        'Total Revenue (KES)', 'Profit (KES)', 'Cash Sales (KES)',
        'M-Pesa Sales (KES)', 'Bank Sales (KES)', 'Customers',
        'Sold By', 'Notes'
    ],
    'employee_data': [
        'Employee Name', 'Role', 'Work Location', 'Start Date',
        'Salary (KES/month)', 'Payment Method', 'Phone', 'Status', 'Notes'
    ],
    'payments_data': [
        'Month', 'Employee', 'Amount (KES)', 'Date Paid',
        'Payment Method', 'Status'
    ],
    'egg_data': [
        'Date', 'Day', 'Month', 'Trays', 'Eggs per Tray',
        'Total Eggs', 'Sellable', 'Cracked', 'Quality', 'Notes'
    ],
    'egg_sales_data': [
        'Date', 'Trays Sold', 'Price per Tray (KES)',
        'Total Revenue (KES)', 'Customer', 'Payment Status', 'Notes'
    ],
    'fruit_data': [
        'Date', 'Fruit Type', 'Quantity', 'Unit', 'Quality',
        'Harvested By', 'Selling Price (KES/kg)', 'Total Revenue (KES)', 'Notes'
    ],
    'chick_data': [
        'Hatch Date', 'Breed', '# Hatched', 'Survival',
        'Survival %', 'Location', 'Vaccinated', 'Notes'
    ],
    'pet_feed_data': [
        'Date', 'Pet Type', 'Meal', 'Food Type', 'Quantity (kg)',
        'Cost per kg (KES)', 'Total Cost (KES)', 'Fed By',
        'Animals Fed', 'Notes'
    ],
    'mortality_data': [
        'Date', 'Species', 'Age', 'Lost', 'Reason',
        'Action Taken', 'Prevention', 'Severity'
    ],
    'cat_menu_data': [
        'Date', 'Day', 'Meal', 'Ingredients', 'Portion (cups)',
        'Schedule', 'Cats Fed', 'Notes'
    ],
    'finance_data': [
        'Date', 'Source/Expense', 'Type', 'Amount (KES)',
        'Category', 'Payment Method', 'Notes'
    ],
    'feed_data': [
        'Feed Type', 'Quantity (kg)', 'Cost per kg (KES)',
        'Supplier', 'Last Ordered', 'Reorder Level (kg)', 'Status'
    ],
    'tasks_data': [
        'Date', 'Task', 'Priority', 'Assigned To',
        'Status', 'Category', 'Notes'
    ],
    'weather_data': [
        'Date', 'Temperature (C)', 'Humidity (%)', 'Weather',
        'Rainfall (mm)', 'Notes'
    ],
}

# Friendly sheet names for Excel exports (no emoji, max 31 chars)
SHEET_NAMES = {
    'inventory_data':       'Livestock Inventory',
    'orchard_data':         'Orange Orchard',
    'orange_harvest_data':  'Orange Harvest Log',
    'cereal_data':          'Cereal Shop Sales',
    'cereal_inv_data':      'Cereal Inventory',
    'cereal_daily_data':    'Cereal Daily Sales',
    'egg_data':             'Egg Production',
    'egg_sales_data':       'Egg Sales',
    'employee_data':        'Employees',
    'payments_data':        'Employee Payments',
    'fruit_data':           'Fruit Production',
    'chick_data':           'Hatched Chicks',
    'pet_feed_data':        'Pet Feeding Log',
    'mortality_data':       'Mortality Tracker',
    'cat_menu_data':        'Cat Meal Menu',
    'finance_data':         'Farm Finance',
    'feed_data':            'Feed Inventory',
    'tasks_data':           'Daily Tasks',
    'weather_data':         'Weather Log',
}


def order_columns(session_key, df):
    """Reorder a DataFrame to the canonical column order."""
    if session_key not in COLUMN_ORDER:
        return df
    desired = COLUMN_ORDER[session_key]
    present = [c for c in desired if c in df.columns]
    extra = [c for c in df.columns if c not in present]
    return df[present + extra]


def export_excel(session_key, df):
    """Create a formatted Excel file in memory.
    
    Returns (bytes, filename).
    """
    from io import BytesIO
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    ordered = order_columns(session_key, df)
    # Sort by Date descending if Date column exists
    if 'Date' in ordered.columns:
        ordered = ordered.sort_values('Date', ascending=False).reset_index(drop=True)
    sheet_name = SHEET_NAMES.get(session_key, session_key)[:31]
    
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        ordered.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        
        # Header styling
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        
        for col_idx in range(1, len(ordered.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        
        # Auto-fit column widths
        for col_idx, col_name in enumerate(ordered.columns, 1):
            max_len = len(str(col_name))
            for row_idx in range(2, len(ordered) + 2):
                cell_val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                max_len = max(max_len, len(cell_val))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 35)
        
        # Data cell styling
        data_align = Alignment(vertical='center')
        alt_fill = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')
        for row_idx in range(2, len(ordered) + 2):
            for col_idx in range(1, len(ordered.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.alignment = data_align
                cell.border = thin_border
                if row_idx % 2 == 0:
                    cell.fill = alt_fill
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    filename = f"liz_farm_{session_key.replace('_data', '').replace('_', '-')}.xlsx"
    return buffer.getvalue(), filename
